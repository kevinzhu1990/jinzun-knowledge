#!/usr/bin/env python3
import json, re, collections, math
from pathlib import Path
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parents[1]
PRODUCT_JSON=ROOT/'outputs/product_quiz/金尊产品知识库题库.json'
ROLE_JSON=ROOT/'outputs/role_quiz/岗位学习考核题库.json'
SOURCE_XLSX=Path('/Users/liangyanmei/Downloads/金尊产品知识库_Excel版---d9395a19-d2c1-488b-99a9-030f909c05ab.xlsx')
OUT=ROOT/'outputs/quiz_logic_audit.json'

retired={
  '1532','1535','1538','1541','1544','1609','1615','1619','1620','2007','2100','2233','2312','2313','2315','2346','2371','2372','2391','2392','2402','2459','2463','2467','2468','2472','2476','2477','2478','2480','2481','2482','2483','2491','2502','2503','2511','2515','2539','2540','2575','1539','1548','1607','1916','2123','2196','2197','2198','2232','2232A','2393','2450','2451','2479','1701彩盒','2097','1183','1656','1690','1701','2071','2077','2091','2098','2201','2202','2217','2220','2221','2331','2345','2421','2429','2431','2532','2551','气泡膜','会员卡','2560','2277+2431ZX一盒装'
}
product_line_overrides={
  '2231':'糕点礼盒类','2179':'曲奇/饼干类','1918':'曲奇/饼干类','2223':'月饼-礼盒','2587':'糕点礼盒类','2367':'糕点礼盒类','2369':'糕点礼盒类','2370':'糕点礼盒类','2576':'糕点类','2577':'糕点类','2578':'糕点类','2582ZX三盒装':'包装/耗材'
}
skip_kp={('2295','保质期')}

# Read questions
product_q=json.loads(PRODUCT_JSON.read_text(encoding='utf-8'))
role_q=json.loads(ROLE_JSON.read_text(encoding='utf-8'))
all_q=product_q+role_q

# Read source workbook with formulas resolved if cached; raw otherwise fallback
wb=load_workbook(SOURCE_XLSX, data_only=True, read_only=True)
ws=wb['产品总表']
rows=list(ws.iter_rows(values_only=True))
headers=[str(x).strip() if x is not None else '' for x in rows[0]]
def get(row,name):
    try: v=row[headers.index(name)]
    except ValueError: return ''
    if v is None: return ''
    s=str(v).strip()
    return '' if s.lower()=='nan' else s

def norm(v):
    return re.sub(r'；\s*；','；',re.sub(r'\s+',' ',re.sub(r'\r?\n+','；',str(v or '')))).strip()
def shelf(v):
    t=norm(v); m=re.search(r'\d+',t)
    if not m: return t
    n=int(m.group()); days=n if '天' in t else n*30 if '月' in t else n
    return {30:'30天',45:'45天',90:'90天/3个月',180:'180天/6个月',270:'270天/9个月',300:'300天/10个月',360:'360天/12个月'}.get(days,t)
def is_retired(code):
    return any(str(code).startswith(r) for r in retired)
source=[]
for row in rows[1:]:
    d={
      'code':norm(get(row,'货号')), 'name':norm(get(row,'产品名称')), 'category':norm(get(row,'一级分类')), 'productLine':norm(get(row,'产品线')),
      'cartonSpec':norm(get(row,'箱规')), 'contents':norm(get(row,'内配/口味')), 'netWeight':norm(get(row,'净重')),
      'shelfLife':shelf(get(row,'保质期')), 'sizeOrBox':norm(get(row,'尺寸/外箱')), 'unit':norm(get(row,'单位')),
    }
    if d['code'] and d['name'] and d['category'] and not is_retired(d['code']):
        if d['code'] in product_line_overrides: d['productLine']=product_line_overrides[d['code']]
        if 'ZX' in d['code']: d['productLine']='包装/耗材'
        source.append(d)
by_code={d['code']:d for d in source}

issues=[]
def issue(level,typ,q,detail,expected=None,actual=None):
    issues.append({'level':level,'type':typ,'id':q.get('id'),'bank':q.get('bank'),'code':q.get('code'),'knowledgePoint':q.get('knowledgePoint'),'question':q.get('question'),'detail':detail,'expected':expected,'actual':actual})

def options(q): return {k:(q.get('option'+k) or '').strip() for k in 'ABCD' if (q.get('option'+k) or '').strip()}
field_by_kp={'产品线':'productLine','箱规':'cartonSpec','产品名称':'name','内配/口味':'contents','克重/净重':'netWeight','保质期':'shelfLife','尺寸/外箱':'sizeOrBox','单位':'unit'}

# generic structural
for q in all_q:
    opts=options(q); ans=q.get('answer')
    if ans not in 'ABCD' or ans not in opts:
        issue('error','答案字母无效',q,'answer 不在现有选项中',actual=ans)
    if ans in opts and q.get('answerText') and opts[ans] != str(q.get('answerText')).strip():
        # 图片选择题的 answerText 是图片路径，选项是图片A-D，属于正常
        if q.get('knowledgePoint') != '看货号选图片':
            issue('error','答案内容与选项不一致',q,'answerText 与正确选项文字不一致',expected=opts[ans],actual=q.get('answerText'))
    rev=collections.defaultdict(list)
    for k,v in opts.items(): rev[v].append(k)
    for text,letters in rev.items():
        if len(letters)>1: issue('error','重复选项',q,f'{letters} 都是 {text}')
    if len(opts)<4 and q.get('type') not in ['判断题']:
        issue('warn','选项不足',q,f'非判断题只有 {len(opts)} 个选项')

# product source consistency
seen_expected=set()
for q in product_q:
    code=q.get('code','')
    src=by_code.get(code)
    if not src:
        issue('error','货号不在源表/已退役',q,'题目货号不在当前有效产品源表中')
        continue
    if q.get('productName') != src['name']:
        issue('error','题目产品名与源表不一致',q,'productName 与源表产品名称不同',expected=src['name'],actual=q.get('productName'))
    kp=q.get('knowledgePoint')
    field=field_by_kp.get(kp)
    if field:
        seen_expected.add((code,kp))
        exp=src[field]
        if q.get('answerText') != exp:
            issue('error','答案与源表不一致',q,f'{kp} 答案不等于源表字段 {field}',expected=exp,actual=q.get('answerText'))
        if q.get('explanation') and exp and exp not in q.get('explanation'):
            issue('warn','解析未包含正确答案',q,'解析中未直接包含源表答案',expected=exp,actual=q.get('explanation'))
    elif kp=='看图片选货号':
        if q.get('answerText') != code:
            issue('error','图片题答案货号不一致',q,'看图片选货号 answerText 应等于本题 code',expected=code,actual=q.get('answerText'))
    elif kp=='看货号选图片':
        if code not in str(q.get('answerText')):
            issue('error','图片选择题答案图片不一致',q,'看货号选图片 answerText 应包含本题 code',expected=code,actual=q.get('answerText'))
    elif kp=='运营活动':
        exp=src['productLine'] or src['category']
        if q.get('answerText') != exp:
            issue('error','场景题答案与源表产品线不一致',q,'运营活动答案应等于产品线/分类',expected=exp,actual=q.get('answerText'))

# expected missing base questions for source rows when enough fields exist: only warn for core kps that should exist
present={(q.get('code'),q.get('knowledgePoint')) for q in product_q}
for d in source:
    expected=[('产品线','productLine'),('箱规','cartonSpec'),('产品名称','name')]
    if d['category']=='包装/耗材': expected += [('尺寸/外箱','sizeOrBox'),('单位','unit')]
    else: expected += [('内配/口味','contents'),('克重/净重','netWeight'),('保质期','shelfLife')]
    for kp,field in expected:
        if (d['code'],kp) in skip_kp: continue
        if d[field] and (d['code'],kp) not in present:
            issues.append({'level':'warn','type':'应有题目缺失','id':'','bank':'','code':d['code'],'knowledgePoint':kp,'question':'','detail':f'源表有 {field}={d[field]}，但题库未生成该知识点题','expected':d[field],'actual':''})

summary=collections.Counter((i['level'],i['type']) for i in issues)
report={'counts':{'product_questions':len(product_q),'role_questions':len(role_q),'total_questions':len(all_q),'source_active_products':len(source)},'summary':{f'{k[0]}:{k[1]}':v for k,v in summary.items()},'issues':issues}
OUT.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding='utf-8')
print(OUT)
print(json.dumps(report['counts'], ensure_ascii=False))
print(json.dumps(report['summary'], ensure_ascii=False, indent=2))
for i in issues[:80]: print(json.dumps(i,ensure_ascii=False))
